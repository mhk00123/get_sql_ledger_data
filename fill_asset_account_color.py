# 填色
import openpyxl
from openpyxl.styles import PatternFill

def fill_asset_account_color():
    file_name = '會計科目表.xlsx'
    sheet_name = "會計科目表"

    wb = openpyxl.load_workbook(file_name)
    wb_name = sheet_name

    print(wb_name)

    # 選定表
    wb_sheet_select = wb[wb_name]
    print(wb_sheet_select)

    # solid = 實色填充
    fill_color = PatternFill('solid', fgColor="FFBB02")

    rows = wb_sheet_select.max_row
    print(rows)

    check_list = [1, 21,22,231,232,24,251,252,261,262,263,264,271,272,281,291,3,4,5,6,7,821,822,831,832,891]

    for i in range(2, rows):
        cell_select = wb_sheet_select.cell(row = i, column=1)
        if (int(cell_select.value) in check_list):
            for j in range(1, wb_sheet_select.max_column+1):
                wb_sheet_select.cell(row=i, column=j).fill = fill_color
            print("found")
        i = i+1
        
    wb.save(file_name)
    
def fill_trial_balance_color():
    file_name = "試算表.xlsx"
    sheet_name = "試算表"

    wb = openpyxl.load_workbook(file_name)
    wb_name = sheet_name

    # 選定表
    wb_sheet_select = wb[wb_name]
    print(wb_sheet_select)

    # solid = 實色填充
    fill_color = PatternFill('solid', fgColor="FFBB02")

    rows = wb_sheet_select.max_row
     
    for i in range(2, rows):
        cell_select = wb_sheet_select.cell(row = i, column=1)
        if (cell_select.value == None):
            for j in range(1, wb_sheet_select.max_column+1):
                wb_sheet_select.cell(row=i, column=j).fill = fill_color
            print("found")
        i = i+1
    
    fill_color = PatternFill('solid', fgColor="8DB4E2")
    for j in range(1, wb_sheet_select.max_column+1):
            wb_sheet_select.cell(row=rows, column=j).fill = fill_color
    
    wb.save(file_name)
    
def fill_income_statement_color():
    file_name = "損益表.xlsx"
    sheet_name = "損益表"

    wb = openpyxl.load_workbook(file_name)
    wb_name = sheet_name

    # 選定表
    wb_sheet_select = wb[wb_name]
    print(wb_sheet_select)

    # solid = 實色填充
    rows = wb_sheet_select.max_row
    fill_color = PatternFill('solid', fgColor="8DB4E2")
    for j in range(1, wb_sheet_select.max_column+1):
            wb_sheet_select.cell(row=rows, column=j).fill = fill_color
    
    wb.save(file_name)
    
if __name__ == '__main__':
    fill_income_statement_color()